"""
Build a SMALL multi-tab xlsx of the deterministic field analysis,
suitable for uploading to Google Drive as a native Google Sheet.
ZERO AI. Reads the same SQLite data as analyze-fields.py.
"""
import sqlite3, json, os, statistics
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(__file__), "..", "data", "app.db")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "meta-ads-field-analysis.xlsx")

con = sqlite3.connect(DB)
con.row_factory = sqlite3.Row

def parse_list(val):
    if not val:
        return []
    try:
        v = json.loads(val)
        return v if isinstance(v, list) else [v]
    except Exception:
        return [val]

rows = con.execute("""
    SELECT a.*, c.name AS comp,
           an.angle, an.angle_secondary, an.brand_voice, an.text_density,
           an.primary_conversion_goal, an.themes, an.pain_points, an.benefits,
           an.analysis_failed_at, an.id AS analysis_id
    FROM ads a JOIN competitors c ON c.id=a.competitor_id
    LEFT JOIN ad_analyses an ON an.ad_id=a.id
""").fetchall()
con.close()

COMPS = sorted(set(r["comp"] for r in rows))
analyzed = [r for r in rows if r["analysis_id"] and not r["analysis_failed_at"]]

wb = Workbook()
wb.remove(wb.active)

HFILL = PatternFill("solid", fgColor="2F3640")
HFONT = Font(bold=True, color="FFFFFF")
TITLEFONT = Font(bold=True, size=12, color="6C5CE7")

def new_sheet(name):
    ws = wb.create_sheet(title=name[:31])
    ws.append_row = 1
    return ws

def write_block(ws, title, headers, data_rows):
    r = ws.append_row
    c = ws.cell(row=r, column=1, value=title)
    c.font = TITLEFONT
    r += 1
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font = HFONT
        cell.fill = HFILL
    r += 1
    for dr in data_rows:
        for ci, v in enumerate(dr, 1):
            ws.cell(row=r, column=ci, value=v)
        r += 1
    ws.append_row = r + 1  # blank spacer row

def derive(r, key):
    if key == "status":
        return "Live" if r["is_active"] else "Paused"
    return (r[key] or "(blank)") if isinstance(r[key], (str, type(None))) else r[key]

def cat_block(ws, title, source, keyfn):
    overall = Counter()
    by_comp = {c: Counter() for c in COMPS}
    for r in source:
        v = keyfn(r)
        v = v.strip() if isinstance(v, str) else v
        if v in (None, ""):
            v = "(blank)"
        overall[v] += 1
        by_comp[r["comp"]][v] += 1
    total = sum(overall.values())
    headers = ["Value", "Total", "%"] + COMPS
    data = []
    for val, n in overall.most_common():
        data.append([val, n, round(100*n/total, 1)] + [by_comp[c].get(val, 0) for c in COMPS])
    data.append(["TOTAL", total, 100.0] + [sum(by_comp[c].values()) for c in COMPS])
    write_block(ws, title, headers, data)

def list_block(ws, title, source, key, top=20):
    overall = Counter()
    n_with = 0
    for r in source:
        items = [str(x).strip().lower() for x in parse_list(r[key]) if str(x).strip()]
        if items:
            n_with += 1
        overall.update(items)
    headers = ["Rank", "Value", "Count", "% of contributing ads"]
    data = [[i, v, n, round(100*n/n_with, 1)] for i, (v, n) in enumerate(overall.most_common(top), 1)]
    write_block(ws, f"{title}  (n_with={n_with}, distinct={len(overall)})", headers, data)

def num_block(ws, title, key, buckets):
    vals = [r[key] for r in rows if r[key] is not None]
    nulls = sum(1 for r in rows if r[key] is None)
    headers = ["Stat", "Value"]
    stats = [["count (non-null)", len(vals)], ["null/blank", nulls],
             ["min", min(vals)], ["median", round(statistics.median(vals), 1)],
             ["mean", round(statistics.mean(vals), 1)], ["max", max(vals)]]
    write_block(ws, title, headers, stats)
    bdata = []
    for label, lo, hi in buckets:
        n = sum(1 for v in vals if lo <= v <= hi)
        bdata.append([label, n, round(100*n/len(vals), 1)])
    write_block(ws, f"{title} — buckets", ["Bucket", "Count", "%"], bdata)

# ---- Tab 1: Overview ----
ws = new_sheet("Overview")
cc = Counter(r["comp"] for r in rows)
write_block(ws, "Ad coverage", ["Competitor", "Ads"], [[c, n] for c, n in cc.most_common()] + [["TOTAL", len(rows)]])
af = sum(1 for r in rows if r["analysis_failed_at"])
no = sum(1 for r in rows if not r["analysis_id"])
write_block(ws, "AI-analysis coverage", ["Status", "Ads"],
            [["analyzed OK", len(analyzed)], ["failed (stub)", af], ["never analyzed", no]])

# ---- Tab 2: Status & format (raw) ----
ws = new_sheet("Status & Format")
cat_block(ws, "Live vs Paused (raw Meta)", rows, lambda r: derive(r, "status"))
cat_block(ws, "Media type (raw Meta)", rows, lambda r: r["media_type"])
cat_block(ws, "Display format (raw Meta)", rows, lambda r: r["display_format"])
cat_block(ws, "CTA button (raw Meta)", rows, lambda r: r["cta_label"])

# ---- Tab 3: Reach & longevity ----
ws = new_sheet("Reach & Longevity")
list_block(ws, "Placements (raw Meta)", rows, "placements", top=10)
list_block(ws, "Countries recorded (raw Meta)", rows, "countries", top=20)
num_block(ws, "Days active (run length, raw Meta)", "days_active",
          [("0-7", 0, 7), ("8-30", 8, 30), ("31-90", 31, 90),
           ("91-180", 91, 180), ("181-365", 181, 365), ("365+", 366, 10**9)])
num_block(ws, "Collation count (creative reuse, raw Meta)", "collation_count",
          [("1", 1, 1), ("2-5", 2, 5), ("6-10", 6, 10), ("11-25", 11, 25), ("26+", 26, 10**9)])

# ---- Tab 4: Messaging (AI) ----
ws = new_sheet("Messaging (AI)")
cat_block(ws, "Primary angle (AI)", analyzed, lambda r: r["angle"])
cat_block(ws, "Secondary angle (AI)", analyzed, lambda r: r["angle_secondary"])
cat_block(ws, "Brand voice (AI)", analyzed, lambda r: r["brand_voice"])
cat_block(ws, "Text density (AI)", analyzed, lambda r: r["text_density"])
cat_block(ws, "Conversion goal (CTA-derived)", analyzed, lambda r: r["primary_conversion_goal"])

# ---- Tab 5: Pain / Benefits / Themes (AI) ----
ws = new_sheet("Pain-Benefit-Theme (AI)")
list_block(ws, "Top pain points (AI)", analyzed, "pain_points", top=20)
list_block(ws, "Top benefits (AI)", analyzed, "benefits", top=20)
list_block(ws, "Top themes (AI)", analyzed, "themes", top=20)

# widths
for ws in wb.worksheets:
    ws.column_dimensions["A"].width = 42
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14

wb.save(OUT)
print("saved", OUT, os.path.getsize(OUT), "bytes")
