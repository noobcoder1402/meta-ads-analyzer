import sqlite3, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

DB = os.path.join(os.path.dirname(__file__), "..", "data", "app.db")
OUT = os.path.join(os.path.dirname(__file__), "..", "data", "meta-ads-export.xlsx")

con = sqlite3.connect(DB)
con.row_factory = sqlite3.Row

# Columns: (header, sql expression / key, is_json_array)
COLS = [
    ("Ad Library URL", "__url__", False),
    ("Library ID", "library_id", False),
    ("Status", "__status__", False),
    ("Days Active", "days_active", False),
    ("Media Type", "media_type", False),
    ("Display Format", "display_format", False),
    ("CTA Button", "cta_label", False),
    ("Caption", "caption", False),
    ("Title", "title", False),
    ("Landing URL", "landing_url", False),
    ("Collation Count", "collation_count", False),
    ("Contains AI Media", "__ai__", False),
    ("Placements", "__placements__", False),
    ("First Seen", "first_seen_at", False),
    ("--- AI ANALYSIS ---", "__sep__", False),
    ("Hook", "hook", False),
    ("Angle", "angle", False),
    ("Secondary Angle", "angle_secondary", False),
    ("Subject", "subject", False),
    ("Visual Summary", "visual_summary", False),
    ("Themes", "themes", True),
    ("Pain Points", "pain_points", True),
    ("Benefits", "benefits", True),
    ("Target Persona", "target_persona", False),
    ("Emotional Tone", "emotional_tone", False),
    ("Brand Voice", "brand_voice", False),
    ("Text Density", "text_density", False),
    ("Dominant Colors", "dominant_colors", True),
    ("Analysis Failed?", "__failed__", False),
]

def join_json(val):
    if not val:
        return ""
    try:
        arr = json.loads(val)
        if isinstance(arr, list):
            return " | ".join(str(x) for x in arr)
        return str(arr)
    except Exception:
        return str(val)

def join_placements(val):
    if not val:
        return ""
    try:
        arr = json.loads(val)
        if isinstance(arr, list):
            return ", ".join(str(x) for x in arr)
        return str(arr)
    except Exception:
        return str(val)

competitors = con.execute(
    "SELECT id, name FROM competitors WHERE id IN (SELECT DISTINCT competitor_id FROM ads) ORDER BY name"
).fetchall()

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill("solid", fgColor="2F3640")
header_font = Font(bold=True, color="FFFFFF")
sep_fill = PatternFill("solid", fgColor="6C5CE7")

total = 0
for comp in competitors:
    rows = con.execute(
        """
        SELECT a.*, an.hook, an.angle, an.angle_secondary, an.visual_summary,
               an.dominant_colors, an.text_density, an.subject, an.themes,
               an.pain_points, an.benefits, an.target_persona, an.emotional_tone,
               an.brand_voice, an.analysis_failed_at
        FROM ads a
        LEFT JOIN ad_analyses an ON an.ad_id = a.id
        WHERE a.competitor_id = ?
        ORDER BY a.is_active DESC, a.days_active DESC
        """,
        (comp["id"],),
    ).fetchall()

    # Sheet name max 31 chars, no special chars
    sheet_name = comp["name"].replace("/", "-")[:31]
    ws = wb.create_sheet(title=sheet_name)

    # header row
    for ci, (header, _, _) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font = header_font
        c.fill = sep_fill if header.startswith("---") else header_fill
        c.alignment = Alignment(vertical="center")

    for ri, r in enumerate(rows, start=2):
        for ci, (header, key, is_json) in enumerate(COLS, start=1):
            if key == "__url__":
                v = f"https://www.facebook.com/ads/library/?id={r['library_id']}" if r["library_id"] else ""
            elif key == "__status__":
                v = "Live" if r["is_active"] else "Paused"
            elif key == "__ai__":
                v = "Yes" if r["contains_ai_media"] else "No"
            elif key == "__placements__":
                v = join_placements(r["placements"])
            elif key == "__sep__":
                v = ""
            elif key == "__failed__":
                v = "FAILED" if r["analysis_failed_at"] else ""
            elif is_json:
                v = join_json(r[key])
            else:
                v = r[key] if r[key] is not None else ""
            ws.cell(row=ri, column=ci, value=v)
        total += 1

    # column widths
    widths = {
        "Ad Library URL": 42, "Caption": 60, "Title": 30, "Landing URL": 40,
        "Visual Summary": 50, "Themes": 35, "Pain Points": 40, "Benefits": 40,
        "Target Persona": 50, "Hook": 35, "Library ID": 18,
    }
    for ci, (header, _, _) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(header, 16)
    ws.freeze_panes = "A2"
    print(f"{comp['name']}: {len(rows)} ads")

wb.save(OUT)
print(f"\nTotal: {total} ads -> {OUT}")
con.close()
